"""Excel全シートの最新データをDBに同期する。
- products / inventory / checkout を Excel から再取り込み
- ringi / todos は保持
"""
import openpyxl
import sqlite3
import os
from datetime import datetime, date

EXCEL_PATH = r'C:\Users\acuma\Dropbox\AI活用\総務資料\事務用品持出明細表.xlsx'
DB_PATH = os.path.join(os.path.dirname(__file__), 'soumu.db')


def to_int(v, default=0):
    if v is None or v == '':
        return default
    try:
        return int(v)
    except (ValueError, TypeError):
        return default


def to_str(v):
    return '' if v is None else str(v)


def sync():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    # --- 商品一覧 ---
    ws = wb['商品一覧']
    c.execute('DELETE FROM products')
    p_count = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        code = to_int(row[0], None)
        name = row[1]
        if code is None or not name:
            continue
        c.execute('INSERT OR REPLACE INTO products (code, name, site, formal_name) VALUES (?,?,?,?)',
                  (code, name, to_str(row[2]), to_str(row[3])))
        p_count += 1

    # --- 在庫チェック表 ---
    ws = wb['在庫チェック表(発注担当者用)']
    c.execute('DELETE FROM inventory')
    i_count = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        name = row[1]
        no = to_int(row[2], None)
        if not name or no is None:
            continue
        stock = to_int(row[3])
        order_qty = to_int(row[4], 10)
        remaining = to_int(row[5])
        judgment = to_str(row[6])
        supply = to_int(row[7])
        usage = to_int(row[8])
        threshold = to_int(row[9])
        note = to_str(row[10])
        c.execute('''INSERT INTO inventory
            (product_name, no, stock, order_qty, remaining, order_judgment, supply_qty, usage_qty, order_threshold, note)
            VALUES (?,?,?,?,?,?,?,?,?,?)''',
            (name, no, stock, order_qty, remaining, judgment, supply, usage, threshold, note))
        i_count += 1

    # --- 持出明細表 ---
    ws = wb['持出明細表']
    c.execute('DELETE FROM checkout')
    co_count = 0
    current_year = 2026
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        year_v = row[1]
        if year_v:
            current_year = to_int(year_v, current_year)
        date_v = row[2]
        if not date_v:
            continue
        if isinstance(date_v, datetime):
            date_str = date_v.strftime('%Y-%m-%d')
        elif isinstance(date_v, date):
            date_str = date_v.strftime('%Y-%m-%d')
        else:
            date_str = str(date_v)[:10]
        code = to_int(row[3], None)
        if code is None:
            continue
        product_name = to_str(row[4])
        supply = to_int(row[5])
        usage = to_int(row[6])
        user_name = to_str(row[7])
        note = to_str(row[8])
        c.execute('''INSERT INTO checkout
            (year, date, code, product_name, supply, usage, user_name, note)
            VALUES (?,?,?,?,?,?,?,?)''',
            (current_year, date_str, code, product_name, supply, usage, user_name, note))
        co_count += 1

    conn.commit()

    totals = {
        'products': c.execute('SELECT COUNT(*) FROM products').fetchone()[0],
        'inventory': c.execute('SELECT COUNT(*) FROM inventory').fetchone()[0],
        'checkout': c.execute('SELECT COUNT(*) FROM checkout').fetchone()[0],
        'ringi (preserved)': c.execute('SELECT COUNT(*) FROM ringi').fetchone()[0],
        'todos (preserved)': c.execute('SELECT COUNT(*) FROM todos').fetchone()[0],
    }
    conn.close()

    print(f'Imported: products={p_count}, inventory={i_count}, checkout={co_count}')
    print('DB totals:')
    for k, v in totals.items():
        print(f'  {k}: {v}')


if __name__ == '__main__':
    sync()
